__author__ = 'Gabrielle Martin-Fortier'

from logs.patrouilleur import Patrouilleur
from logs.agent import Agent
from logs.exceptions import PieceInexistante, ErreurDeplacement
from logs.log import Log
import ast
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import os

class Quart:
    """Quart de travail de la sûreté, contenant des patrouilleurs ainsi que leurs affectations de la journée.

    Attributes:
        nb_patrouilleurs (int): Le nombre de patrouilleurs mobiles
        id_lieutenant (agent) : l'identité et les attributs du lieutenant
        id_204 (Patrouilleur) : l'identité et les attributs de l'agent 204
        id_205 (Patrouilleur) : l'identité et les attributs de l'agent 205
        id_206 (Patrouilleur) : l'identité et les attributs de l'agent 205
        id_207 (Patrouilleur) : l'identité et les attributs de l'agent 207
        dernier_pat_log (Patrouilleur) : Le dernier patrouilleur à avoir eu un log
    """

    def __init__(self, jour_nuit=None, id_lieutenant=Agent("", "", ""), couleur_lieutenant="chartreuse", nb_patrouilleurs=1, id_204=None, id_205=None, id_206=None, id_207=None):
        """
        Args:
            id_lieutenant
            nb_patrouilleurs (int) : Le nombre de patrouilleurs dans le quart
            id_204=None (Patrouilleur) :
            id_205=None (Patrouilleur)
            id_206=None (Patrouilleur)
            id_207=None (Patrouilleur)
        """

        self.nb_patrouilleurs = nb_patrouilleurs
        self.id_lieutenant = id_lieutenant
        self.id_patrouilleurs = {
            "204": id_204,
            "205": id_205,
            "206": id_206,
            "207": id_207
        }
        self.couleur_lieutenant = couleur_lieutenant
        self.dernier_pat_log = []
        self.nom_quart = datetime.datetime.now().strftime("%B-%d-%Y %Hh%M")
        self.jour_nuit = jour_nuit

    def effacer_dernier_log(self):
        for pat in self.dernier_pat_log[0]:
            self.id_patrouilleurs[pat].effacer_dernier_log()

    def convertir_en_chaine(self):
        """Retourne une chaîne de caractères où chaque log est écrite sur une ligne distincte.

        Returns:
            (str): La chaîne de caractères construite.

        """
        liste_logs = []
        for log in self.dernier_pat_log:
            liste_logs.append(','.join(log))
        derniers_logs = '-'.join(liste_logs)
        chaine = "{};{};{};{}\n".format(self.nb_patrouilleurs, self.couleur_lieutenant,
                                        self.id_lieutenant.convertir_en_chaine(),
                                        derniers_logs)

        for patrouilleur in self.id_patrouilleurs.values():
            if patrouilleur is not None:
                chaine += patrouilleur.convertir_en_chaine()

        return chaine

    def charger_dune_chaine(self, chaine):
        """Vide le damier actuel et le remplit avec les informations reçues dans une chaîne de caractères. Le format de
        la chaîne est expliqué dans la documentation de la méthode convertir_en_chaine.

        Args:
            chaine (str): La chaîne de caractères.
        """
        for vieux_pats in self.id_patrouilleurs:
            self.id_patrouilleurs[vieux_pats] = None
        numero_de_ligne = 0
        patrouilleur = False
        for information_quart in chaine.split("\n"):
            if information_quart in ["204", "205", "206", "207"]:
                numero_de_ligne = 1
                poste = information_quart
                patrouilleur = True
            if information_quart != "":
                if numero_de_ligne == 0:
                    nb_patrouilleurs, self.couleur_lieutenant, nom_lieutenant, titre_lieutenant, indicatif_lieutenant, \
                    self.dernier_pat_log = information_quart.split(";")
                    self.nb_patrouilleurs = int(nb_patrouilleurs)
                elif patrouilleur:
                    if numero_de_ligne == 2:
                        nom, titre, indicatif, couleur1, couleur2 = information_quart.split(";")
                        self.id_patrouilleurs[poste] = \
                            Patrouilleur(Agent(nom, titre, indicatif), poste, [couleur1, couleur2])
                    elif numero_de_ligne != 1:
                        nouveau_log = Log()
                        nouveau_log.charger_dune_chaine(information_quart)
                        self.id_patrouilleurs[poste].logs.insert(0, nouveau_log)
            numero_de_ligne += 1

        self.id_lieutenant = Agent(nom_lieutenant, titre_lieutenant, indicatif_lieutenant)
        for pat in self.id_patrouilleurs:
            if self.id_patrouilleurs[pat] is not None:
                self.id_patrouilleurs[pat].position = self.id_patrouilleurs[pat].logs[0]
                self.id_patrouilleurs[pat].logs.pop(0)

        self.dernier_pat_log = self.dernier_pat_log.split("-")
        new_list = []
        for item in self.dernier_pat_log:
            new_list.append(item.split(","))
        self.dernier_pat_log = new_list

    def sauvegarder(self, nom_fichier):
        """
        Args:
            nom_fichier (str): Le nom du fichier où sauvegarder.
        """
        if not os.path.exists("Quarts format texte"):
            os.mkdir("Quarts format texte")

        nom_fichier = os.getcwd() + "/Quarts format texte/" + nom_fichier
        with open(nom_fichier, "w") as f:
            f.writelines(self.convertir_en_chaine())
        f.close()

    def cellule_titre(self, ws, ligne, colonne, contenu, fontsize):
        ws.cell(row=ligne, column=colonne, value=contenu).font = Font(size=fontsize, bold=True)
        ws.cell(row=ligne, column=colonne).alignment = Alignment(horizontal='center')

    def sauvegarder_excel(self, nom_fichier):
        wb = Workbook()
        sheet = wb.get_sheet_by_name('Sheet')
        wb.remove_sheet(sheet)
        for pat in self.id_patrouilleurs:
            if self.id_patrouilleurs[pat] is not None:
                ws = wb.create_sheet(pat)
                columns = ['A', 'B', 'C', 'D', 'E', 'F']
                for c in columns:
                    ws.column_dimensions[c].width = 30

                self.cellule_titre(ws, 1, 1, int(pat), 20)
                self.cellule_titre(ws, 1, 2, self.id_patrouilleurs[pat].agent.titre, 20)
                self.cellule_titre(ws, 1, 3, self.id_patrouilleurs[pat].agent.nom, 20)
                self.cellule_titre(ws, 1, 4, int(self.id_patrouilleurs[pat].agent.indicatif), 20)

                self.cellule_titre(ws, 3, 1, "Heure début", 14)
                self.cellule_titre(ws, 3, 2, "Heure fin", 14)
                self.cellule_titre(ws, 3, 3, "Terminal ou extérieur", 14)
                self.cellule_titre(ws, 3, 4, "Côté ville ou piste", 14)
                self.cellule_titre(ws, 3, 5, "Log", 14)
                self.cellule_titre(ws, 3, 6, "Note", 14)

                self.id_patrouilleurs[pat].logs.insert(0, self.id_patrouilleurs[pat].position)
                for x in range(0, len(self.id_patrouilleurs[pat].logs)):
                    logs_reversed = self.id_patrouilleurs[pat].logs[::-1]
                    ws.cell(row=x+4, column=1, value=logs_reversed[x].heure_debut)
                    ws.cell(row=x+4, column=2, value=logs_reversed[x].heure_fin)
                    ws.cell(row=x+4, column=3, value=logs_reversed[x].terminal_exterieur)
                    ws.cell(row=x+4, column=4, value=logs_reversed[x].cote)
                    ws.cell(row=x+4, column=5, value=logs_reversed[x].log)
                    ws.cell(row=x+4, column=6, value=logs_reversed[x].note)

        #Ici on place les feuilles de travail dans un ordre croissant selon les postes des patrouilleurs
        new_order = sorted(wb.sheetnames)
        new_sheets = []
        for sheet in new_order:
            new_sheets.append(wb.get_sheet_by_name(sheet))
        wb._sheets = new_sheets

        nom = nom_fichier
        wb.save(nom)
